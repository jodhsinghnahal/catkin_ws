# This is a program to create the alert table in the
# xnet-params database and fill it in with data for the FEX4K

###############################################################################
#                               Includes
###############################################################################

import sqlite3
from openpyxl import load_workbook
from RvcDefs import *
import Fex4kAlerts
import PgDatabase
from MakeParamDatabase import fnGetModelId
from Rvc2Mqtt import ZRPM_BMS_MODELS

###############################################################################
#                             Constant Data
###############################################################################

# Database file name
DB_FILE = 'xnet_params.db'

# table creation
ALERTS_TABLE = """create table alerts (
                   model_id integer,
                   code integer,
                   condition text,
                   mode text,
                   action text,
                   unique (model_id, code),
                   foreign key (model_id) references model (id) on delete cascade
                );"""

RVC_ALERT_MAPPING_TABLE = """create table rvc_alert_mapping (
                               model_id integer,
                               alert_type text,
                               code integer,
                               spn text,
                               fmi text,
                               unique (model_id, alert_type, code),
                               foreign key (model_id) references model (id) on delete cascade
                             );"""


MB_ALERT_MAPPING_TABLE = """create table mb_alert_mapping (
                              model_id integer,
                              code integer,
                              register integer,
                              bitmask integer,
                              alert_type text,
                              value integer,
                              unique (model_id, code),
                              foreign key (model_id) references model (id) on delete cascade
                            );"""

INSERT_ALERT_SQL = """insert into alerts (model_id, code, condition, mode, action)
                      values (?, ?, ?, ?, ?);"""

FXCC_SPREADSHEET = 'docs/FxcProSupport.xlsx'
FXCC_ALERTS_SHEET = 'Alerts'

# Content for the FEX4K's alert table, mined from its user manual
ALERT = [
    (1, "AC Output Under Voltage",
     "Any mode",
     "Clear the fault and attempt to restart. Contact customer service if problem persists."),
    (2, "AC Output Over Voltage",
     "Any mode",
     "Clear the fault and attempt to restart. Contact customer service if problem persists."),
    (9, "BMS Node Missing for CVCC mode",
     "Shore (AC) mode",
     "This fault triggers when the charger stage is set to CVCC but there is no BMS node found on the network."),
    (17, "Relay(s) Welded",
     "Shore (AC) mode",
     "The AC L1 transfer relay is bad or an AC source was wired directly to the AC output. Disconnect the inverter’s output wiring. If error continues, have unit serviced."),
    (19, "Relay(s) Welded",
     "Shore (AC) mode",
     "The AC L1 transfer relay is bad or an AC source was wired directly to the AC output. Disconnect the inverter’s output wiring. If error continues, have unit serviced."),
    (18, "Relay(s) Welded",
     "Shore (AC) mode",
     "The AC L2 transfer relay is bad or an AC source was wired directly to the AC output. Disconnect the inverter’s output wiring. If error continues, have unit serviced."),
    (20, "Relay(s) Welded",
     "Shore (AC) mode",
     "The AC L2 transfer relay is bad or an AC source was wired directly to the AC output. Disconnect the inverter’s output wiring. If error continues, have unit serviced."),
    (21, "Relay(s) Welded",
     "Shore (AC) mode",
     "The AC L1L2 transfer relay is bad or an AC source was wired directly to the AC output. Disconnect the inverter’s output wiring. If error continues, have unit serviced."),
    (22, "Relay(s) Welded",
     "Shore (AC) mode",
     "The AC L1 transfer relay is bad or an AC source was wired directly to the AC output. Disconnect the inverter’s output wiring. If error continues, have unit serviced."),
    (41, "Auxiliary Power Supply Under Voltage Shutdown",
     "Any mode",
     "Clear the fault and attempt restart. If problem persists, call customer service."),
    (42, "Auxiliary Power Supply Over Voltage Shutdown",
     "Any mode",
     "Clear the fault and attempt restart. If problem persists, call customer service."),
    (44, "Battery Over Temperature Shutdown",
     "Any mode",
     "Check battery voltage and battery cable connections. Stop charging, if necessary. Check for excessive ambient temperature and adequate ventilation in the battery compartment."),
    (45, "Capacitor Over Temperature Shutdown",
     "Any mode",
     "Clear the fault and attempt restart. Ensure adequate ventilation. Reduce AC loads."),
    (46, "Controller Fault",
     "Any mode",
     "Service required."),
    (47, "DC Under Immediate Voltage Shutdown",
     "Inverter (Battery) mode",
     "Check battery status and recharge if necessary. Check for proper DC cable sizing. Check for loose connections and tighten if necessary."),
    (48, "DC Under Voltage Shutdown",
     "Inverter (Battery) mode",
     "Check battery status and recharge if necessary. Check for proper DC cable sizing. Check for loose connections and tighten if necessary."),
    (49, "DC Over Voltage Shutdown",
     "Inverter (Battery) mode",
     "Check for external charging sources, such as a PV charger and an overvoltage alternator. Disconnect, if necessary."),
    (51, "EEPROM Error",
     "Any mode",
     "No action. Clear fault and resume operating or configuring the unit. If the fault persists, have the unit serviced."),
    (52, "EEPROM Error",
     "Any mode",
     "No action. Clear fault and resume operating or configuring the unit. If the fault persists, have the unit serviced."),
    (53, "EEPROM Error",
     "Any mode",
     "No action. Clear fault and resume operating or configuring the unit. If the fault persists, have the unit serviced."),
    (54, "EEPROM Error",
     "Any mode",
     "No action. Clear fault and resume operating or configuring the unit. If the fault persists, have the unit serviced."),
    (55, "EEPROM Error",
     "Any mode",
     "No action. Clear fault and resume operating or configuring the unit. If the fault persists, have the unit serviced."),
    (56, "EEPROM Error",
     "Any mode",
     "No action. Clear fault and resume operating or configuring the unit. If the fault persists, have the unit serviced."),
    (57, "FET1 Over Temperature Shutdown",
     "Any mode",
     "Reduce the loads connected to the AC outlet of the unit. Check that the ventilation grille is not blocked. Check for ambient temperature and move the unit ot a cooler location whenever possible. Check the fan for any obstruction and remove it."),
    (58, "FET2 Over Temperature Shutdown",
     "Any mode",
     "Reduce the loads connected to the AC outlet of the unit. Check that the ventilation grille is not blocked. Check for ambient temperature and move the unit ot a cooler location whenever possible. Check the fan for any obstruction and remove it."),
    (60, "Unit Fault",
     "Any mode",
     "Clear faults/warnings and restart. If error persists, contact customer service."),
    (61, "Unit Fault",
     "Any mode",
     "Clear faults/warnings and restart. If error persists, contact customer service."),
    (62, "Unit Fault",
     "Any mode",
     "Clear faults/warnings and restart. If error persists, contact customer service."),
    (63, "AC Overload",
     "Any mode",
     "Check for loads above the inverter’s capacity, turn off some loads if necessary. Power down and restart the unit to clear the manual fault."),
    (64, "AC Overload Line 1",
     "Any mode",
     "Check for loads above the inverter’s capacity, turn off some loads if necessary. Power down and restart the unit to clear the manual fault."),
    (65, "AC Overload Line 2",
     "Any mode",
     "Check for loads above the inverter’s capacity, turn off some loads if necessary. Power down and restart the unit to clear the manual fault."),
    (66, "System Configuration Fault",
     "Multi-unit configuration settings",
     "Ensure each unit has a unique Device Number, and that associations and physical connections have been configured correctly."),
    (67, "Watchdog Error",
     "Any mode",
     "Clear faults/warnings and restart. If error persists, contact customer service."),
    (68, "Transformer Over Temperature",
     "Any mode",
     "Reduce the loads connected to the AC outlet of the unit. Check that the ventilation grille is not blocked. Check for ambient temperature and move the unit ot a cooler location whenever possible. Check the fan for any obstruction and remove it."),
    (69, "External Sync Failed",
     "Multi-unit configuration",
     "Check connections and cable on external AC sync port. In a single- inverter system, nothing must be plugged into the AC sync port. Clear fault and try again. If these steps fail, the unit requires service."),
    (71, "Battery Discharge Over Current",
     "Inverter (Battery) mode",
     "There is an excessive load on the Li-Ion battery. (The fault applies only to Li-Ion batteries.) Change the default threshold of the max battery discharge current limit or reduce the load."),
    (72, "External AC Contactor Malfunction",
     "Shore (AC) mode",
     "Check why the AC contactor has failed. Check for fusing of coil, wiring and connections. Verify that the AC contactor has power."),
    (73, "DC-DC Converter FETA Over Temperature",
     "DC-DC Converter",
"Reduce the loads connected to the DC load of the unit. Check that the ventilation grille is not blocked. Check for ambient temperature and move the unit ot a cooler location whenever possible. Check the fan for any obstruction and remove it."),
    (74, "DC-DC Converter FETB Over Temperature",
     "DC-DC Converter",
"Reduce the loads connected to the DC load of the unit. Check that the ventilation grille is not blocked. Check for ambient temperature and move the unit ot a cooler location whenever possible. Check the fan for any obstruction and remove it."),
     (75, "DC-DC Converter Under Voltage",
      "DC-DC Converter",
      "Remove the loads connected to the DC-DC converter and try again. Remove the AC loads of the unit."),
     (76, "DC-DC Converter Over Voltage",
      "DC-DC Converter",
      "Check for external charging sources, such as a PV charger or an overvoltage alternator. Disconnect if necessary"),
     (77, "DC-DC Converter Over Current",
      "DC-DC Converter",
      "Remove the loads connected to the DC-DC converter and try again. Remove the AC loads of the unit."),
     (87, "Fan fatigue warning",
      "Any mode",
      "If there is no issue with the fan/s, disconnect the unit from its DC and AC power sources, then reconnect, and then restart the unit. Perform Step 8: Testing Your Installation. If error detection persists, contact customer service."),
     (88, "Fan fatigue warning",
      "Any mode",
      "If there is no issue with the fan/s, disconnect the unit from its DC and AC power sources, then reconnect, and then restart the unit. Perform Step 8: Testing Your Installation. If error detection persists, contact customer service."),
     (89, "Fan lock warning",
      "Any mode",
      "If there is no issue with the fan/s, disconnect the unit from its DC and AC power sources, then reconnect, and then restart the unit. Perform Step 8: Testing Your Installation. If error detection persists, contact customer service."),
     (90, "Fan lock warning",
      "Any mode",
      "If there is no issue with the fan/s, disconnect the unit from its DC and AC power sources, then reconnect, and then restart the unit. Perform Step 8: Testing Your Installation. If error detection persists, contact customer service."),
     (94, "Remote Power Off",
      "Any mode",
      "No action required. The unit stops inverting or charging immediately, and shuts down after five seconds. If the unit is configured as a master, it signals other network devices to also shut down."),
     (95, "Equalization Abort",
      "Shore (AC) mode",
      "Equalization terminated abnormally because of interrupted AC input. Wait until AC input returns to in-tolerance condition."),
     (96, "Cannot Equalize",
      "Shore (AC) mode",
      "Change battery type if your batteries should be equalized. Gel or AGM batteries should not be equalized. AC input is not qualified or the charge setting is not adequate."),
     (97, "Battery Sensor Short",
      "Any mode",
      "Replace battery temperature sensor.")
    ]

# Verify fault codes in database with fault codes in the DiagMsg1 lookup
def fnTestFex4k(db):
    cursor = db.cursor()
    cursor.execute("create temporary table fex4k (code integer, condition text);")
    cursor.close()

    codes = set()
    for tup in Fex4kAlerts.atzFltLookup:
        codes.add(tup[0])

    for tup in Fex4kAlerts.atzWrnLookup:
        codes.add(tup[0])

    for code in codes:
        cursor = db.cursor()
        cursor.execute("insert into fex4k (code, condition) values (?, ?);",
                       (code, Fex4kAlerts.fnGetAlertString(code)))
        cursor.close()

    # What are we missing from the alerts table?
    cursor = db.cursor()
    cursor.execute("""select fex4k.code, fex4k.condition
                      from fex4k left join alerts
                      on fex4k.code = alerts.code
                      where alerts.model_id = 1 and alerts.code is null
                      order by fex4k.code""")
    print("Missing from alerts table:")
    while True:
        try:
            code, cond = cursor.fetchone()
        except:
            break

        print(f"{code}, {cond}")

    cursor.close()

    # What are we missing from the Rvc2Mqtt faults and warnings lookup?
    cursor = db.cursor()
    cursor.execute("""select alerts.code, alerts.condition
                      from alerts left join fex4k
                      on fex4k.code = alerts.code
                      where alerts.model_id = 1 and fex4k.code is null
                      order by alerts.code""")
    print("Missing from rvc2mqtt lookup table:")
    while True:
        try:
            code, cond = cursor.fetchone()
        except:
            break

        print(f"{code}, {cond}")

    cursor.close()


    # What codes mis-match?
    print("Mis-matched condition strings:")
    cursor = db.cursor()
    cursor.execute("""select alerts.code, alerts.condition, fex4k.condition
                      from alerts inner join fex4k
                      on fex4k.code = alerts.code
                      where alerts.model_id = 1 and fex4k.condition != alerts.condition
                      order by alerts.code""")

    while True:
        try:
            code, alerts_cond, fex4k_cond = cursor.fetchone()
        except:
            break

        print(f"{code}, {alerts_cond}, {fex4k_cond}")

    cursor.close()

    # What codes match exactly?
    print("Matching condition strings:")
    cursor = db.cursor()
    cursor.execute("""select alerts.code, alerts.condition
                      from alerts inner join fex4k
                      on fex4k.code = alerts.code
                      where alerts.model_id = 1 and fex4k.condition = alerts.condition
                      order by alerts.code""")

    while True:
        try:
            code, cond = cursor.fetchone()
        except:
            break

        print(f"{code}, {cond}")

    cursor.close()

# Convert an RV-C Fault Mode Indicator (FMI) number into text
def fnGetFmiName(fmi):
    fmi_dict = PgDatabase.ENUMLOOKUP[9419]
    return fmi_dict.get(fmi, f"?{fmi}?")

# Convert an RV-C Service Point Number (SPN) into text
def fnGetInvSpnName(spn_msb, spn_isb, spn_lsb):
    # Is this a global SPN?
    if spn_msb == 0:
        spn_dict = PgDatabase.GLOBAL_SPNS
        try:
            spn_name = spn_dict[(spn_msb, spn_isb, spn_lsb)]
        except KeyError:
            spn = (spn_isb << 3) + spn_lsb
            spn_name = f"?{spn}?"

    else:
        # This is a device-specific SPN for the inverter or inverter/charger
        spn_dict = PgDatabase.SPN_MAPPING[66]
        try:
            spn_name = spn_dict[(spn_msb, spn_lsb)]
            if spn_isb not in [0, 255]:
                spn_name += f' {spn_isb}'

        except KeyError:
            spn_name = f"?{spn_msb}/{spn_isb}/{spn_lsb}?"

    return spn_name

# Read in Lithionics BMS alert data from spreadsheet
# and insert into the alert table
def fnInsertBmsData(db, spreadsheetname, model_id):
    wb = load_workbook(spreadsheetname, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        # print(row)
        field, code, mnem, severity, condition, action = row
        if condition:
            cursor = db.cursor()
            cursor.execute(INSERT_ALERT_SQL, (model_id,
                                              code,
                                              condition,
                                              None,
                                              action))
            cursor.close()


# Read in Modbus Solar Charge Controller alert data from spreadsheet
def fnInsertMbSccData(db, spreadsheetname, model_id):
    wb = load_workbook(spreadsheetname, data_only=True)
    ws = wb['Alerts']

    for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        code = row[-2]
        condition = row[-1]
        addr = row[1]  # This column is expressed in hex
        bitmask = row[4]
        warnval = row[7]
        faultval = row[8]
        if addr is not None:
            addr = int(str(addr), 16)
        if faultval is not None:
            alert_val = faultval
            alert_type = 'fault'
        else:
            alert_val = warnval
            alert_type = 'warning'

        if code is not None:
            # The code is a hex number in the spreadsheet
            code = int(str(code), 16)
            bitmask = int(bitmask, 16)
            #print(f"Inserting SCC code {code:05x}, {condition}")
            cursor = db.cursor()
            cursor.execute(INSERT_ALERT_SQL, (model_id,
                                              code,
                                              condition,
                                              None,
                                              None))
            cursor.close()

            # Insert modbus details into mb_alert_mapping
            cursor = db.cursor()
            cursor.execute("""insert into mb_alert_mapping
                              (model_id, code, register, bitmask, alert_type, value)
                              values (?, ?, ?, ?, ?, ?);""",
                           (model_id, code, addr, bitmask, alert_type, alert_val))
            cursor.close()

# Read in RV-C Solar Charge Controller alert data from spreadsheet
def fnInsertRvcSccData(db, spreadsheetname, model_id):
    wb = load_workbook(spreadsheetname, data_only=True)
    ws = wb['Alerts']

    for row in ws.iter_rows(min_row=2, max_col=17, values_only=True):
        warnval = row[7]
        faultval = row[8]
        code = row[9]
        condition = row[10]
        spn_name = row[11]
        spn_msb = row[12]
        spn_isb = row[13]
        spn_lsb = row[14]
        fmi_name = row[15]
        fmi = row[16]
        if faultval is not None:
            alert_val = faultval
            alert_type = 'fault'
        else:
            alert_val = warnval
            alert_type = 'warning'

        if code is not None and spn_name is not None:
            # The code is a hex number in the spreadsheet
            code = int(str(code), 16)

            cursor = db.cursor()
            cursor.execute(INSERT_ALERT_SQL, (model_id,
                                              code,
                                              condition,
                                              None,
                                              None))
            cursor.close()

            # Insert RV-C details into rvc_alert_mapping
            cursor = db.cursor()
            cursor.execute("""insert into rvc_alert_mapping
                              (model_id, alert_type, code,
                               spn, fmi)
                              values (?, ?, ?, ?, ?);""",
                           (model_id, alert_type, code, spn_name, fmi_name))
            cursor.close()


# Add entries into the rvc_alert_mapping table for the FEX4K
def fnAddRvcMapping(db, model_id, alert_type, lookup_table):
    sql = """insert into rvc_alert_mapping
             (model_id, alert_type, code,
              spn, fmi)
             values (?, ?, ?, ?, ?);"""
    for code, spn_msb, spn_isb, spn_lsb, fmi in lookup_table:
        # Convert FMI number into text
        fmi_name = fnGetFmiName(fmi)

        # Convert SPN into text
        spn_name = fnGetInvSpnName(spn_msb, spn_isb, spn_lsb)

        cursor = db.cursor()
        try:
            cursor.execute(sql, (model_id, alert_type, code, spn_name, fmi_name))
        except sqlite3.IntegrityError:
            print(f"For model {model_id}, {alert_type} {code} is non-unique")
            raise

        cursor.close()

# Add entries into the rvc_alert_mapping table for the FXC Pro (FXCC)
def fnAddFxccMapping(db, model_id, spreadsheet, sheet):
    wb = load_workbook(spreadsheet)
    ws = wb[sheet]

    alert_sql = """insert into alerts
                   (model_id, code, condition, action)
                   values (?, ?, ?, ?);"""

    rvc_sql = """insert into rvc_alert_mapping
                 (model_id, alert_type, code,
                  spn, fmi)
                 values (?, ?, ?, ?, ?);"""

    for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        code, flag, condition, alert_type, spn_name, fmi_name, spn_msb, spn_isb, spn_lsb, fmi_num, action = row
        if code and condition:
            # Make an entry into alerts
            cursor = db.cursor()
            try:
                cursor.execute(alert_sql, (model_id, code, condition, action))
            except sqlite3.IntegrityError:
                print(f"For FXC model {model_id}, code {code} is non-unique")

            cursor.close()

            # Make an entry into rvc_alert_mapping
            if spn_msb != 0 and spn_isb:
                # An instance was specified for the SPN, append that number to the lookup text
                spn_name = f"{spn_name} {spn_isb}"

            cursor = db.cursor()
            try:
                cursor.execute(rvc_sql, (model_id, alert_type, code, spn_name, fmi_name))
            except sqlite3.IntegrityError:
                print(f"For FXC model {model_id}, {alert_type} {code} is non-unique")
                raise

            cursor.close()


# The main program starts here
def fnMain():
    # Create a connection to the new database
    db = sqlite3.Connection(DB_FILE)

    # Drop the alerts-related tables if they exist already
    for tablename in ['alerts', 'rvc_alert_mapping', 'mb_alert_mapping']:
        cursor = db.cursor()
        cursor.execute("drop table if exists %s;" % tablename)
        cursor.close()

    # Create the tables
    for table_creation in [ALERTS_TABLE, RVC_ALERT_MAPPING_TABLE, MB_ALERT_MAPPING_TABLE]:
        cursor = db.cursor()
        cursor.execute(table_creation)
        cursor.close()

    # Fill the table with data for the FEX4K
    model_id = fnGetModelId(db, name='FEX_RVC_4048_120_60')
    for code, condition, mode, action in ALERT:
        cursor = db.cursor()
        cursor.execute(INSERT_ALERT_SQL, (model_id, code, condition, mode, action))
        cursor.close()

    # Fill out the RVC SPN/FMI lookups for the FEX4K
    model_id = fnGetModelId(db, name='FEX_RVC_4048_120_60')
    fnAddRvcMapping(db, model_id, 'fault', Fex4kAlerts.atzFltLookup)
    fnAddRvcMapping(db, model_id, 'warning', Fex4kAlerts.atzWrnLookup)

    # Fill out the RVC SPN/FMI lookups and descriptions for the Freedom XC and X series
    cursor = db.cursor()
    cursor.execute("select id from model where friendly_name like 'Freedom X%'")
    while True:
        try:
            model_id = cursor.fetchone()[0]
        except:
            break

        fnAddFxccMapping(db, model_id, FXCC_SPREADSHEET, FXCC_ALERTS_SHEET)

    cursor.close()

    # Read in Lithionics BMS alert data from docs/LithionicsFlags.xlsx
    # and insert into the alert table
    model_id = fnGetModelId(db, friendly_name="NeverDie")
    fnInsertBmsData(db, 'docs/LithionicsFlags.xlsx', model_id)

    # Read in ZeroRPM BMS alert data from docs/ZeroRpmFlags.xlsx
    # and insert into the alert table
    for model_name in ZRPM_BMS_MODELS:
        model_id = fnGetModelId(db, name=model_name)
        fnInsertBmsData(db, 'docs/ZeroRpmFlags.xlsx', model_id)

    # Read in Modbus Solar Charge Controller data from docs/SccParams.xlsx
    for model_name in ['DR3210N', 'LS GM3024N']:
        model_id = fnGetModelId(db, name=model_name)
        fnInsertMbSccData(db, 'docs/SccParams.xlsx', model_id)

    # Read in RV-C Solar Charge Controller data from the same spreadsheet
    model_id = fnGetModelId(db, name='MPPT_RVC_30')
    fnInsertRvcSccData(db, 'docs/SccParams.xlsx', model_id)

    # For now, copy the same data for the SRNE MPPT 60
    model_id = fnGetModelId(db, name='MC4860N15')
    fnInsertRvcSccData(db, 'docs/SccParams.xlsx', model_id)

    # Add in system alerts for the RVC network
    model_id = fnGetModelId(db, name='rvc')
    alert_sql = """insert into alerts
                   (model_id, code, condition, action)
                   values (?, ?, ?, ?);"""
    cursor = db.cursor()
    cursor.execute(alert_sql,
                   (model_id, 1, 'Duplicate device instances found',
                    "Duplicate instances of the same device type were found. This is usually due to devices not being configured properly or being reset to factory defaults. Each device needs to be reconfigured with a unique instance for its specific role within your system. Until then, all settings made to a device will also be applied to other devices with the same type and instance. Proceed with caution and contact your installer for more information."))
    cursor.close()

    db.commit()

    fnTestFex4k(db)

    db.close()


if __name__ == '__main__':
    fnMain()

