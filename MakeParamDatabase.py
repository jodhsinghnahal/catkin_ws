# This is a program to create the parameters database for the Xantrex
# Communications Platform.  It includes setting limits for each supported
# model of connected power device.

###############################################################################
#                               Includes
###############################################################################

import sqlite3
from openpyxl import load_workbook
import re
from PgDesc import clXbDatabase

###############################################################################
#                             Constant Data
###############################################################################

# The source of much of the info on this spreadsheet
SPREADSHEET_NAME = 'docs/RvcParams.xlsx'

# A lookup for which column indicates support for which product family
# Column I indicates parameter support for the FEX4k
# Column J indicates support for the FXC pro
# Column K indicates support for the FX inverters
# Column L indicates support for the Lithonics NeverDie BMS
# Column M indicates support for the ZeroRPM BMS
# Column N indicates support for the RV-C Solar Charge Controller
# Column O indicates support for the SRNE Solar charge Controllers
RVC_SUPPORT_LIST = [(8, 'Freedom EX%'),
                    (9, 'Freedom XC%'),
                    (10, 'Freedom X %'),
                    (11, 'NeverDie'),
                    (12, 'ZeroRPM%'),
                    (13, 'MPPT Charge Controller 30 RVC'),
                    (14, 'MPPT Charge Controller 60')]

SETTINGS_SPREADSHEET = 'docs/Freedom 4048 - Remote Panel Settings.xlsx'
FEX4K_SETTINGS_SHEET = 'Settings Menu Map'

SCC_SPREADSHEET = 'docs/SccParams.xlsx'
SRNE_SPREADSHEET = 'docs/SrneMpptSupport.xlsx'
SETTINGS_SHEET = 'Settings'

# SCC_SETTINGS_SPREADSHEET = 'docs/SCC WebApp Settings_Description 2021-04-26.xlsx'

FXCC_SPREADSHEET = 'docs/FxcProSupport.xlsx'

# Database file name
DB_FILE = 'xnet_params.db'

# Table creation
param_table = """create table if not exists param (
                    id integer primary key autoincrement,
                    name text not null,
                    desc text,
                    label text,
                    type text,
                    access text,
                    units text,
                    enum_id integer,
                    foreign key (enum_id) references enum (id) on delete set null
                 );"""

rvc_support_table = """create table if not exists rvc_support (
                          param_id integer not null,
                          model_id integer not null,
                          sts_pgn text,
                          sts_sig text,
                          cmd_pgn text,
                          cmd_sig text,
                          from_sts_trans_func text,
                          to_cmd_trans_func text,
                          unique (param_id, model_id),
                          foreign key (param_id) references param (id) on delete cascade,
                          foreign key (model_id) references model (id) on delete cascade
                       );"""

rvc_qual_table = """create table rvc_qual (
                       param_id integer not null,
                       model_id integer not null,
                       signal text not null,
                       value text,
                       unique (param_id, model_id, signal)
                       foreign key (param_id) references param(id) on delete cascade,
                       foreign key (model_id) references model (id) on delete cascade
                     );"""

modbus_support = """create table if not exists modbus_support (
                       model_id integer,
                       param_id integer,
                       reg_addr integer,
                       class text default 'holding',
                       registers integer default 1,
                       scale integer default 1,
                       bitmask integer,
                       enum_id integer default null,
                       unique (model_id, param_id),
                       foreign key (model_id) references model (id) on delete cascade,
                       foreign key (param_id) references param (id) on delete cascade,
                       foreign key (enum_id) references modbus_enum (id) on delete set null
                    );"""

modbus_enum = """create table if not exists modbus_enum (
                    id integer primary key autoincrement,
                    name text not null
                 );"""

modbus_enum_values = """create table if not exists modbus_enum_values (
                           enum_id integer,
                           enum_num integer,
                           enum_value text,
                           foreign key (enum_id) references modbus_enum (id) on delete cascade
                        );"""

# TODO: Separate tables for Modbus enums?

model_table = """create table if not exists model (
                    id integer primary key autoincrement,
                    devtype text,
                    name text not null,
                    friendly_name text not null,
                    desc text
                 );"""

model_support = """create table if not exists model_support (
                      model_id integer,
                      param_id integer,
                      default_text text,
                      default_float real,
                      min_float real,
                      max_float real,
                      step_float real,
                      enum_id integer,
                      desc text,
                      caution text,
                      level text default 'basic',
                      write_protect integer default 0,
                      unique (model_id, param_id),
                      foreign key (model_id) references model (id) on delete cascade,
                      foreign key (param_id) references param (id) on delete cascade,
                      foreign key (enum_id) references enum (id) on delete set null
                   );"""

enum_table = """create table if not exists enum (
                   id integer primary key autoincrement,
                   name text not null,
                   desc text
                );"""

enum_values_table = """create table if not exists enum_values (
                          enum_id integer,
                          enum_ord integer,
                          enum_val text,
                          unique (enum_id, enum_ord),
                          foreign key (enum_id) references enum (id) on delete cascade
                       );"""

ALL_TABLES = [param_table, rvc_support_table, rvc_qual_table,
              model_table, model_support,
              modbus_support, modbus_enum, modbus_enum_values,
              enum_table, enum_values_table]

QUAL_MATCHER = re.compile(r"(\w+)\s*==\s*\W?([-_ a-zA-Z0-9]*)\W?")

###############################################################################
#                           Function Definitions
###############################################################################

# Get the model ID, given device type or model name
def fnGetModelId(db, **kwargs):
    # Make up the query, based on keyword arguments
    sql = "select id from model where "
    params = []
    specifiers = []
    for key, val in kwargs.items():
        specifiers.append('%s = ?' % key)
        params.append(val)

    sql += ' and '.join(specifiers)
    sql += ';'

    cursor = db.cursor()
    cursor.execute(sql, params)
    model_id = cursor.fetchone()[0]
    cursor.close()

    return model_id

# Get the parameter ID, given the name or label
def fnGetParamId(db, **kwargs):
    # Make up the query, based on keyword arguments
    sql = "select id from param where "
    params = []
    specifiers = []
    for key, val in kwargs.items():
        specifiers.append('%s = ?' % key)
        params.append(val)

    sql += ' and '.join(specifiers)
    sql += ';'

    cursor = db.cursor()
    cursor.execute(sql, params)
    try:
        param_id = cursor.fetchone()[0]
    except:
        param_id = None
    finally:
        cursor.close()

    return param_id

# Look up the data type, engineering units and enum DDID of a parameter, given PGN.signal
def fnParamInfo(xbdb, pgn_mnem, sig_mnem):
    # PmAssocSts is munged with association type and instance
    # These are all enums, but the specific enum depends on whether this
    # is an AC or DC association
    if pgn_mnem is not None and pgn_mnem.startswith('PmAssocSts'):
        match = re.match(r'^PmAssocSts([AD])', pgn_mnem)
        if match:
            assoc_type = match.group(1)
            if assoc_type == 'A':
                ddid = xbdb.AC_ASSOC_DDID
            else:
                ddid = xbdb.DC_ASSOC_DDID

            return ('enum', None, ddid)

        pgn_mnem = 'PmAssocSts'

    sigdict = xbdb.fnFindSigDesc(pgn_mnem, sig_mnem)
    if sigdict is None:
        return (None, None, None)

    if sigdict['typename'] == 'enum':
        return ('enum', None, sigdict['ddid'])

    elif sigdict['units'] is None:
        return ('int', None, None)

    else:
        return ('float', sigdict['units'], None)

# Populate the parameters and rvc support tables
def fnPopulateParams(db, ws, xbdb):
    enumids = set()

    for row in ws.iter_rows(min_row=3, max_col=16, values_only=True):
        param_name, sts_pgn, sts_sig, qual, cmd_pgn, cmd_sig, from_sts_trans_func, to_cmd_trans_func, fex4k, fxcc, fx, lithionics, zerorpm, scc, srne_scc, label = row
        if param_name:
            # Is the parameter rw, ro or wo?
            if sts_pgn:
                if cmd_pgn:
                    access = 'rw'
                else:
                    access = 'ro'
            else:
                if cmd_pgn:
                    access = 'wo'
                else:
                    # No RV-C support for this parameter
                    continue

            # Create a new entry in the param table if it doesn't exist already
            c1 = db.cursor()
            c1.execute("select id from param where name = ?;", (param_name,))
            try:
                param_id = c1.fetchone()[0]
            except:
                sql = "insert into param (name, access, label) values (?, ?, ?);"
                cursor = db.cursor()
                cursor.execute(sql, (param_name, access, label))
                param_id = cursor.lastrowid
                cursor.close()
            finally:
                c1.close()

            # Insert an entry for rvc_support for every model supported
            for col, pattern in RVC_SUPPORT_LIST:
                supported = row[col]
                #print(f"Support for {pattern} is {supported}")
                if supported:
                    c1 = db.cursor()
                    c1.execute(f"select id from model where friendly_name like '{pattern}';")
                    while True:
                        try:
                            model_id = c1.fetchone()[0]
                        except:
                            break

                        # print(f"Inserting {param_name} for {pattern} (model {model_id})")

                        sql2 = """insert into rvc_support
                                  (param_id, model_id, sts_pgn, sts_sig, cmd_pgn, cmd_sig, from_sts_trans_func, to_cmd_trans_func)
                                  values (?, ?, ?, ?, ?, ?, ?, ?);"""
                        cursor = db.cursor()
                        cursor.execute(sql2, (param_id, model_id, sts_pgn, sts_sig, cmd_pgn, cmd_sig, from_sts_trans_func, to_cmd_trans_func))
                        cursor.close()

                        # Bust the qualifier string into pieces and add entries to the rvc_qual table
                        sql3 = "insert into rvc_qual (param_id, model_id, signal, value) values (?, ?, ?, ?);"
                        if qual:
                            for substr in qual.split('&&'):
                                match = QUAL_MATCHER.search(substr)
                                if match:
                                    sig = match.group(1)
                                    val = match.group(2)
                                    cursor = db.cursor()
                                    cursor.execute(sql3, (param_id, model_id, sig, val))
                                    cursor.close()

                    c1.close()


                    # Get some info about the parameter
                    if sts_pgn is not None:
                        typename, units, enumid = fnParamInfo(xbdb, sts_pgn, sts_sig)
                    elif cmd_pgn is not None:
                        typename, units, enumid = fnParamInfo(xbdb, cmd_pgn, cmd_sig)
                    else:
                        print(f"No info rvc found for {param_name}")
                        typename = units = enumid = None

                    if typename is not None:
                        sql4 = """update param set type = ?, units = ?, enum_id = ? where id = ?;"""
                        cursor = db.cursor()
                        cursor.execute(sql4, (typename, units, enumid, param_id))
                        cursor.close()

                    if enumid is not None:
                        enumids.add(enumid)

    return enumids

# Add entries for model support, based on a spreadsheet column
def fnPopulateModelSupport(db, ws, model_id, col):
    model_name = fnGetModelName(db, model_id)
    for row in ws.iter_rows(min_row=3, values_only=True):
        param_name = row[0]
        supported = row[col]
        if param_name and supported:
            # Look up the id of the parameter name
            cursor = db.cursor()
            cursor.execute("select id, enum_id from param where name = ?;", (param_name,))
            param_id, enum_id = cursor.fetchone()
            cursor.close()

            # Make a custom copy of the default enum list so that it can
            # be edited down to just what the product supports
            if enum_id is not None:
                model_enum = fnCreateCustomEnum(db, enum_id, param_name, model_id)
            else:
                model_enum = None

            # Make a new blank entry into the model support table
            cursor = db.cursor()

            # For the FEX4K, the description is in column Q and the caution is in column R
            if model_name == 'FEX_RVC_4048_120_60':
                desc = row[16]
                caution = row[17]

                # Generally, if a caution is defined then this is an advanced setting
                if caution or param_name in ['ChargerEnable', 'InverterLowBatteryCutoutVoltage']:
                    level = 'advanced'
                else:
                    level = 'basic'

                cursor.execute("insert into model_support (model_id, param_id, enum_id, desc, caution, level) values (?, ?, ?, ?, ?, ?);",
                                   (model_id, param_id, model_enum, desc, caution, level))
            else:
                # Not an FEX4K or SRNE -> insert with no description or caution text
                cursor.execute("insert into model_support (model_id, param_id, enum_id) values (?, ?, ?);",
                            (model_id, param_id, model_enum))

            cursor.close()

# Get the name for a parameter
def fnGetParamName(db, param_id):
    cursor = db.cursor()
    cursor.execute("select name from param where id = ?;", (param_id,))
    param_name = cursor.fetchone()[0]
    cursor.close()

    return param_name

# Get the name for a model
def fnGetModelName(db, model_id):
    cursor = db.cursor()
    cursor.execute("select name from model where id = ?;", (model_id,))
    model_name = cursor.fetchone()[0]
    cursor.close()

    return model_name

# Create an enum that is custom to a model's parameter so that it can
# be reduced to just the values that the product supports
def fnCreateCustomEnum(db, enum_id, param_name, model_id):
    # Get the name of the model so we can give our enum a nice name
    model_name = fnGetModelName(db, model_id)
    enum_name = f"{model_name}_{param_name}"

    # Create a new entry in the enum table
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values (?);", (enum_name,))
    new_id = cursor.lastrowid
    cursor.close()

    # Copy entries from the old id to the new id
    cursor = db.cursor()
    cursor.execute("select enum_ord, enum_val from enum_values where enum_id = ?",
                   (enum_id,))
    while True:
        try:
            enum_ord, enum_val = cursor.fetchone()
        except:
            break

        c2 = db.cursor()
        c2.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
                   (new_id, enum_ord, enum_val))
        c2.close()

    cursor.close()

    return new_id


# Populate the enum tables, based on used enum IDs from the PyRvc database
def fnPopulateEnums(db, xbdb, enumids):
    for enumid in enumids:
        cursor = db.cursor()
        sql = "insert into enum (id, name) values (?, ?);"
        cursor.execute(sql, (enumid, f"DD{enumid}"))
        cursor.close()

        sql2 = "insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);"
        for i, txtval in enumerate(xbdb.fnEnumList(enumid)):
            # Skip meaningless values
            if txtval in ['No Data', 'No Change']:
                continue

            cursor = db.cursor()
            cursor.execute(sql2, (enumid, i, txtval))
            cursor.close()

# Populate the limits for model support
def fnUpdateModelLimits(db, model_id):
    wb = load_workbook(SETTINGS_SPREADSHEET)
    ws = wb[FEX4K_SETTINGS_SHEET]

    update_sql = """update model_support set
                    default_text = ?, default_float = ?,
                    min_float = ?, max_float = ?, step_float = ?
                    where model_id = ? and param_id = ?"""

    for row in ws.iter_rows(min_row=2, values_only=True):
        # The xnet parameter name is in column D
        param_name = row[3]

        if param_name:
            # Get the ID for the name
            param_id = fnGetParamId(db, name=param_name)

            # Setting default, min, max and step are in columns G to J respectively
            default_float, min_float, max_float, step_float = row[6:10]
            default_text = None
            if default_float is not None and type(default_float) == str:
                # This could define an enum default
                default_text = default_float
                default_float = None

            # Don't save non-numeric values in the min, max and step fields
            if min_float is not None and type(min_float) == str:
                print(f"Minimum for {param_name} is not a number: {min_float}")
                min_float = None

            if max_float is not None and type(max_float) == str:
                print(f"Maximum for {param_name} is not a number: {max_float}")
                max_float = None

            if step_float is not None and type(step_float) == str:
                print(f"Step size for {param_name} is not a number: {step_float}")
                step_float = None

            cursor = db.cursor()
            cursor.execute(update_sql,
                           (default_text, default_float,
                            min_float, max_float, step_float,
                            model_id, param_id))
            cursor.close()


# Turn DefaultBatteryTemperature into a customized enum
def fnDefaultBattTempEnum(db):
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values ('DefaultBattTemp');")
    enum_id = cursor.lastrowid
    cursor.close()

    for i, text in enumerate(['Cold (10 deg C)', 'Warm (25 deg C)', 'Hot (40 deg C)']):
        cursor = db.cursor()
        cursor.execute('insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);',
                       (enum_id, i, text))
        cursor.close()

    cursor = db.cursor()
    cursor.execute("""update param set
                      type = 'enum',
                      units = null,
                      enum_id = ?
                      where name = 'DefaultBatteryTemperature';""",
                    (enum_id,))
    cursor.close()

# Turn InverterOutVoltage into a customized enum
def fnInverterOutVoltageEnum(db):
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values ('InverterOutVoltage');")
    enum_id = cursor.lastrowid
    cursor.close()

    for i, text in enumerate(['108 V', '110 V', '120 V']):
        cursor = db.cursor()
        cursor.execute('insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);',
                       (enum_id, i, text))
        cursor.close()

    cursor = db.cursor()
    cursor.execute("""update param set
                      type = 'enum',
                      units = null,
                      enum_id = ?
                      where name = 'InverterOutVoltage';""",
                    (enum_id,))
    cursor.close()

# Turn InverterEnable into an On/Off enum
def fnInverterEnableEnum(db):
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values ('InverterEnable');")
    enum_id = cursor.lastrowid
    cursor.close()

    for i, text in enumerate(['Off', 'On']):
        cursor = db.cursor()
        cursor.execute('insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);',
                       (enum_id, i, text))
        cursor.close()

    cursor = db.cursor()
    cursor.execute("""update param set
                      type = 'enum',
                      units = null,
                      enum_id = ?
                      where name = 'InverterEnable';""",
                    (enum_id,))
    cursor.close()

# Create a custom enum for InvFaultRecovery, that has Auto/Manual as the options
def fnInvFaultRecoveryEnum(db):
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values ('InvFaultRecovery');")
    enum_id = cursor.lastrowid
    cursor.close()

    for i, text in enumerate(['Auto', 'Manual']):
        cursor = db.cursor()
        cursor.execute('insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);',
                       (enum_id, i, text))
        cursor.close()

    cursor = db.cursor()
    cursor.execute("""update param set
                      type = 'enum',
                      units = null,
                      enum_id = ?
                      where name = 'InvFaultRecovery';""",
                    (enum_id,))
    cursor.close()

# LBCO delay timer enum (so we get variable step sizes)
def fnLBCODelayEnum(db):
    fnInverterDelayEnum(db, 'InverterLowBatteryCutoutDelayTime')

# InverterOutPowerLimitTimer (FXC* only) has the same options
def fnPwrLimitTimerEnum(db):
    fnInverterDelayEnum(db, 'InverterOutPowerLimitTimer')

def fnInverterDelayEnum(db, param_name):
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values (?);", (param_name,))
    enum_id = cursor.lastrowid
    cursor.close()

    # From 1 to 19, step size is 1
    enum_ord = 1
    for i in range(1, 20):
        cursor = db.cursor()
        cursor.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
                       (enum_id, enum_ord, f"{i} s"))
        cursor.close()
        enum_ord += 1

    # From 20 to 300, step size is 10
    for i in range(20, 310, 10):
        cursor = db.cursor()
        cursor.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
                       (enum_id, enum_ord, f"{i} s"))
        cursor.close()
        enum_ord += 1

    # Tie into parameter
    # But for all but FEX4K, which has its own range
    cursor = db.cursor()
    cursor.execute("""update param set
                      type = 'enum',
                      units = null,
                      enum_id = ?
                      where name = ?;""",
                   (enum_id, param_name))
    cursor.close()

# PwrSaveTime (FXC* only) Off (0) and 1 to 25 hours
def fnPwrSaveTimeEnum(db):
    cursor = db.cursor()
    cursor.execute("insert into enum (name) values ('PwrSaveTime');")
    enum_id = cursor.lastrowid
    cursor.close()

    # Add 1 to 25 hours first
    enum_ord = 1
    for i in range(1, 26):
        db.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
                   (enum_id, enum_ord, f"{i} h"))
        enum_ord += 1

    # Last option is 'Off'
    db.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
               (enum_id, enum_ord, 'Off'))

    # Tie into parameter
    db.execute("""update param set
                  type = 'enum', enum_id = ?
                  where name = 'PwrSaveTime';""",
               (enum_id,))

# Erase enums for InverterLowBatteryCutoutDelayTime on FEX4K models and
# replace with enums specific for their range (0 to 600 step 5)
def fnReplace4kLbcoDelay(db):
    # Find the enum_id for the FEX4K's LBCO delay
    cursor = db.cursor()
    cursor.execute("""select model_support.enum_id
                      from model_support inner join param
                      on model_support.param_id = param.id
                      inner join model
                      on model_support.model_id = model.id
                      where param.name = 'InverterLowBatteryCutoutDelayTime'
                      and model.friendly_name like 'Freedom EX%';""")
    while True:
        try:
            enum_id = cursor.fetchone()[0]
        except:
            break

        # Delete all entries for this particular enum
        c2 = db.cursor()
        c2.execute("delete from enum_values where enum_id = ?", (enum_id,))
        c2.close()

        # Now replace with new entries (0 to 600, step 5)
        enum_ord = 1
        for val in range(0, 605, 5):
            c3 = db.cursor()
            c3.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
                       (enum_id, enum_ord, f'{val} s'))
            c3.close()
            enum_ord += 1

    cursor.close()

# Erase enum for InverterOutVoltage on Freedom X 2000 230V and replace with enum specific to its
# ratings (220, 230, 240)
def fnReplace230vInverterOutVoltage(db):
    cursor = db.cursor()
    cursor.execute("""select model_support.enum_id
                      from model_support inner join param
                      on model_support.param_id = param.id
                      inner join model
                      on model_support.model_id = model.id
                      where param.name = 'InverterOutVoltage'
                      and model.friendly_name = 'Freedom X 2000 230V';""")
    while True:
        try:
            enum_id = cursor.fetchone()[0]
        except:
            break

        # Delete all entries for this particular enum
        c2 = db.cursor()
        c2.execute("delete from enum_values where enum_id = ?", (enum_id,))
        c2.close()

        # Now replace with new entries (220, 230 and 240)
        enum_ord = 1
        for val in range(220, 250, 10):
            c3 = db.cursor()
            c3.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?);",
                       (enum_id, enum_ord, f'{val} V'))
            c3.close()
            enum_ord += 1

    cursor.close()

# Turn battery temperature compensation coefficient into a float instead
# of an int (RVC database automatically turns this into an int because it
# is a bit field
def fnTempCompConstIsFloat(db):
    cursor = db.cursor()
    cursor.execute("""update param set
                      type = 'float'
                      where name = 'CustomBatteryTemperatureCompensation'
                      or name = 'BattTempConstScc';""")
    # I'd put in units but this can vary beween the RV-C standard (mv/K)
    # and the EPSolar SCC on Modbus (mv/C/2V)
    cursor.close()

# Add support for the solar charge controller
def fnSolarChargeControllerSupport(db, spreadsheet, model_id):
    wb = load_workbook(spreadsheet)
    ws = wb['Params']

    # Go through the parameter lines
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        param_name, label, hexaddr, addr, regclass, regs, units, scale, bitmask = row
        if param_name and hexaddr:
            # Convert the hexadecimal number string into an integer
            try:
                intaddr = int(str(hexaddr), 16)
            except TypeError:
                print(f'hexaddr {hexaddr} is type {type(hexaddr)}')
                raise

            # Convert the hexadecimal bitmask for partial-register parameters
            if bitmask is not None:
                try:
                    bitmask = int(str(bitmask), 16)
                except TypeError:
                    print(f'Invalid bitmask {bitmask}')
                    bitmask = None

            # Is the parameter already in the database?
            param_id = fnGetParamId(db, name=param_name)
            if param_id is None:
                # Add a new entry into the param table, leave the label out for
                # another, more definitive spreadsheet
                param_id = fnAddSolarChargeParam(db, param_name, None, regclass, units)

            else:
                # Update the units and type if not specified yet
                fnUpdateSolarChargeParam(db, param_id, regclass, units)

            # Make entry into modbus_support table
            fnAddModbusReg(db, param_id, model_id, intaddr, regclass, regs, scale, bitmask)

            # Make entry into model_support table if this is a setting
            if regclass in ['holding', 'coil']:
                # We don't know any limits or default info at this stage.
                # We'll add enum info later on if we find any for this parameter
                db.execute("insert into model_support (model_id, param_id) values (?, ?);",
                           (model_id, param_id))

    # Go through the enums
    ws = wb['Enums']
    enum_dict = {}
    settings_dict = {}
    model_name = fnGetModelName(db, model_id)
    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        param_name, notes, mb_value, enum_text, order = row
        if param_name and (mb_value is not None):
            if param_name not in enum_dict:
                # Create a new modbus enum and put it in the dictionary we're creating
                enum_name = f"{model_name}_{param_name}"
                cursor = db.cursor()
                cursor.execute("insert into modbus_enum (name) values (?);", (enum_name,))
                enum_id = cursor.lastrowid
                cursor.close()
                enum_dict[param_name] = enum_id

            else:
                enum_id = enum_dict[param_name]

            # Add entry to the modbus_enum_values table
            cursor = db.cursor()
            cursor.execute("""insert into modbus_enum_values
                              (enum_id, enum_num, enum_value)
                              values (?, ?, ?);""",
                           (enum_id, mb_value, enum_text))
            cursor.close()

        if param_name and order is not None:
            # Add enum to model_support, enum, enum_values if this is a setting
            if param_name not in settings_dict:
                # Create a new enum for model_support
                enum_name = f"{model_name}_{param_name}"
                cursor = db.cursor()
                cursor.execute("insert into enum (name) values (?);", (enum_name,))
                enum_id = cursor.lastrowid
                cursor.close()
                settings_dict[param_name] = enum_id

                # Tie this new enum ID into the model_support table
                cursor = db.cursor()
                param_id = fnGetParamId(db, name=param_name)
                cursor.execute("""update model_support
                                  set enum_id = ?
                                  where model_id = ? and param_id = ?;""",
                               (enum_id, model_id, param_id))
                cursor.close()

            else:
                enum_id = settings_dict[param_name]

            # Add entry to the enum_values table
            cursor = db.cursor()
            cursor.execute("""insert into enum_values
                              (enum_id, enum_ord, enum_val)
                              values (?, ?, ?);""",
                           (enum_id, order, enum_text))
            cursor.close()



    # Tie enums to enum_id in modbus_support table
    for param_name, enum_id in enum_dict.items():
        param_id = fnGetParamId(db, name=param_name)
        cursor = db.cursor()
        cursor.execute("""update modbus_support
                          set enum_id = ?
                          where model_id = ? and param_id = ?;""",
                       (enum_id, model_id, param_id))
        cursor.close()


# Add a new parameter to the parameter table, based on SCC data
def fnAddSolarChargeParam(db, param_name, label, regclass, units):
    # Determine access, based on the register class
    access = {'input' : 'ro', 'holding' : 'rw', 'coil' : 'wo', 'disc' : 'ro'}.get(regclass)

    # Is this a parameter with units or an enum?
    typename = fnGetTypeName(regclass, units)
    if typename == 'enum':
        units = None

    # Add new entry to param table
    cursor = db.cursor()
    cursor.execute("insert into param (name, label, type, access, units) values (?, ?, ?, ?, ?);",
                   (param_name, label, typename, access, units))
    param_id = cursor.lastrowid
    cursor.close()

    # Return the new parameter ID
    return param_id

# Update a parameter entry to specify units and typename if not defined yet
def fnUpdateSolarChargeParam(db, param_id, regclass, units):
    # Is the typename defined yet?
    cursor = db.cursor()
    cursor.execute("select id from param where id = ? and type is null;", (param_id,))
    tup = cursor.fetchone()
    cursor.close()
    
    if tup is not None:
        # Update the entry with the missing typename and units
        typename = fnGetTypeName(regclass, units)
        if typename == 'enum':
            units = None

        cursor = db.cursor()
        cursor.execute("update param set type = ?, units = ? where id = ?",
                       (typename, units, param_id))
        cursor.close()

# Get the typename for the param table, given the Modbus register class and units
def fnGetTypeName(regclass, units):
    if units == 'enum':
        typename = 'enum'

    elif bool(units):
        typename = 'float'

    else:
        typename = 'int'

    return typename
    

# Add a new entry to the modbus_support table
def fnAddModbusReg(db, param_id, model_id, addr, regclass, regs, scale, bitmask):
    # Default number of regs is 1
    if not bool(regs):
        regs = 1

    if not bool(scale):
        scale = 1

    cursor = db.cursor()
    cursor.execute("""insert into modbus_support
                      (model_id, param_id, reg_addr, class, registers, scale, bitmask)
                      values (?, ?, ?, ?, ?, ?, ?);""",
                   (model_id, param_id, addr, regclass, regs, scale, bitmask))
    cursor.close()

# Add setting descriptions and cautions from a spreadsheet
def fnSolarChargeDescriptions(db, spreadsheet, model_id):
    wb = load_workbook(spreadsheet)
    ws = wb['Settings']

    # Go through the parameter lines
    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        label, param_name, desc, level, caution, default_float, min_float, max_float, step_float = row

        default_text = None
        if default_float is not None and type(default_float) == str:
            # This could define an enum default
            default_text = default_float
            default_float = None

        # Find the parameter ID, based on the name
        param_id = fnGetParamId(db, name=param_name)
        if param_id is not None:
            sql = """update model_support
                     set desc = ?, caution = ?, level = ?,
                     default_text = ?, default_float = ?,
                     min_float = ?, max_float = ?, step_float = ?
                     where model_id = ?
                     and param_id = ?;"""
            cursor = db.cursor()
            cursor.execute(sql, (desc, caution, level,
                                 default_text, default_float,
                                 min_float, max_float, step_float,
                                 model_id, param_id))
            cursor.close()

            sql2 = """update param set label = ?
                      where id = ? and label is null;"""
            cursor = db.cursor()
            cursor.execute(sql2, (label, param_id))
            cursor.close()

# Get parameter descriptions and limits for the FXC Pro
def fnGetDescriptions(db, spreadsheet, model_id):
    wb = load_workbook(spreadsheet)
    ws = wb[SETTINGS_SHEET]

    update_sql = """update model_support set
                    default_text = ?, default_float = ?,
                    min_float = ?, max_float = ?, step_float = ?,
                    desc = ?, caution = ?, level = ?
                    where model_id = ? and param_id = ?;"""

    for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
        param_name, label, level, default_float, min_float, max_float, step_float, desc, caution = row
        if param_name and label:
            # Look up the id for the parameter name
            param_id = fnGetParamId(db, name=param_name)
            if param_id is None:
                raise Exception(f"Could not find id for parameter {param_name}")

            default_text = None
            if default_float is not None and type(default_float) == str:
                # This could define an enum default
                default_text = default_float
                default_float = None

            # Don't save non-numeric values in the min, max and step fields
            if min_float is not None and type(min_float) == str:
                print(f"Minimum for {param_name} is not a number: {min_float}")
                min_float = None

            if max_float is not None and type(max_float) == str:
                print(f"Maximum for {param_name} is not a number: {max_float}")
                max_float = None

            if step_float is not None and type(step_float) == str:
                print(f"Step size for {param_name} is not a number: {step_float}")
                step_float = None

            if level is None:
                level = 'basic'
            elif level not in ['basic', 'advanced']:
                raise Exception(f"FXCC setting level {level} for {param_name} is invalid")

            cursor = db.cursor()
            cursor.execute(update_sql,
                           (default_text, default_float,
                            min_float, max_float, step_float,
                            desc, caution, level,
                            model_id, param_id))
            cursor.close()

            # Look up the label in the database. Update if null, flag if different
            cursor = db.cursor()
            cursor.execute("select label from param where id = ?;", (param_id,))
            param_label = cursor.fetchone()[0]
            cursor.close()

            if param_label is None:
                cursor = db.cursor()
                cursor.execute("update param set label = ? where id = ?;", (label, param_id))
                cursor.close()

            elif param_label != label:
                print(f"FXCC label for {param_name}, {label} != {param_label}")

# Hand-edits for PPN parameters, as they are not in the RVC database
def fnZeroRpmEdits(db):
    # Give each PPN parameter a type
    ppn_types = {'CanAddr' : 'int',
                 'BattInst' : 'int',
                 'NumPacks' : 'int',
                 'BankInst' : 'int',
                 'ReserveLimit' : 'float',
                 'ProdId' : 'int',
                 'RatedCapacity' : 'float',
                 'CellsInSeries' : 'int',
                 'CellsInParallel' : 'int',
                 'BmsSerialNum' : 'string',
                 'BpcSerialNum' : 'string',
                 'BpcHwVer' : 'string',
                 'BlueToothCfg' : 'enum',
                 'WiFiCfg' : 'enum'}
    sql = "update param set type = ? where name = ?;"
    for p_name, p_type in ppn_types.items():
        db.execute(sql, (p_type, p_name))

    # Give limits to the editable int or float parameters
    # Tuples contain default, min, max and step
    ppn_limits = {'CanAddr' : (70.0, 0.0, 70.0, 1.0),
                  'BattInst' : (1.0, 1.0, 16.0, 1.0),
                  'NumPacks' : (1.0, 1.0, 16.0, 1.0),
                  'BankInst' : (1.0, 1.0, 2.0, 1.0),
                  'ReserveLimit' : (10.0, 0.0, 100.0, 1.0)}
    sql1 = "select id from param where name = ?"
    sql2 = """update model_support
    set default_float = ?, min_float = ?, max_float = ?, step_float = ?
    where param_id = ?"""
    for p_name, valtup in ppn_limits.items():
        # Look up the parameter ID
        p_id = fnGetParamId(db, name=p_name)

        p_default, p_min, p_max, p_step = valtup
        db.execute(sql2, (p_default, p_min, p_max, p_step, p_id))

    # Insert enums for PmPpnWriteBlueToothCfg PmPpnWriteWiFiCfg
    ppn_enums = {'BlueToothCfg' : [(0, 'Classic'),
                                   (1, 'BLE')],
                 'WiFiCfg' : [(0, 'Internal'),
                              (1, 'Dongle')]}
    for p_name, enum_list in ppn_enums.items():
        # Create an enum for this parameter
        db.execute("insert into enum (name) values (?);", (p_name,))

        # Get the ID of the new enum
        cursor = db.cursor()
        cursor.execute("select id from enum where name = ?", (p_name,))
        enum_id = cursor.fetchone()[0]

        # Create the enum values
        for enum_ord, enum_val in enum_list:
            db.execute("insert into enum_values (enum_id, enum_ord, enum_val) values (?, ?, ?)",
                       (enum_id, enum_ord, enum_val))

        # Insert the enum into model_support
        p_id = fnGetParamId(db, name=p_name)
        db.execute("update model_support set enum_id = ? where param_id = ?",
                   (enum_id, p_id))

        # Insert the enum id into param
        db.execute("update param set enum_id = ? where name = ?",
                   (enum_id, p_name))

# The main program starts here
def fnMain():
    # Create a connection to the new database
    db = sqlite3.Connection(DB_FILE)

    # Drop all existing tables in this database that we will be creating later
    for table in ['param', 'model', 'rvc_support', 'model_support', 'rvc_qual',
                  'modbus_support', 'modbus_enum', 'modbus_enum_values',
                  'enum_values', 'enum']:
        db.execute(f'drop table if exists {table};')

    # Create the tables
    for table_cmd in ALL_TABLES:
        try:
            db.execute(table_cmd)
        except:
            print(table_cmd)
            raise

    # Open the spreadsheet
    wb = load_workbook(SPREADSHEET_NAME)
    ws = wb.active

    xbdb = clXbDatabase()

    # Populate the models table
    for i, modtup in enumerate([('invchg', 'Freedom EX 4048', 'FEX_RVC_4048_120_60'),
                                ('invchg', 'Freedom XC Pro 2000', 'FX-8182010'),
                                ('invchg', 'Freedom XC Pro 3000', 'FX-8183011'),
                                ('invchg', 'Freedom XC 1000', 'FX-8171050'),
                                ('invchg', 'Freedom XC 2000', 'FX-8172080'),
                                ('invchg', 'Freedom XC 2000 230V', 'FX-817208012'),

                                # The following models are just inverters, but we're calling them
                                # inverter/chargers because they occupy the same spot in the sytem diagram
                                ('invchg', 'Freedom X 1200', 'FX-8061212'),
                                ('invchg', 'Freedom X 1000', 'FX-8171000'),
                                ('invchg', 'Freedom X 2000', 'FX-8172000'),
                                ('invchg', 'Freedom X 3000', 'FX-8173000'),
                                ('invchg', 'Freedom X 2000 24V', 'FX-817200021'),
                                ('invchg', 'Freedom X 2000 230V', 'FX-817200012'),

                                ('bms', 'NeverDie', '8'),
                                ('bms', 'ZeroRPM 310Ah 12V', '0884-0310-12'),
                                ('bms', 'ZeroRPM 410Ah 12V', '0884-0410-12'),
                                ('bms', 'ZeroRPM 310Ah 12V UL', '0884-0310-12-01'),
                                ('bms', 'ZeroRPM 410Ah 12V UL', '0884-0410-12-01'),
                                ('bms', 'ZeroRPM 205Ah 24V', '0884-0205-24'),
                                ('bms', 'ZeroRPM 100Ah 51V', '0884-0100-51'),
                                ('scc', 'MPPT Charge Controller 30 RVC', 'MPPT_RVC_30'),
                                ('scc', 'MPPT Charge Controller 30', 'DR3210N'),
                                ('scc', 'PWM Charge Controller 30', 'LS GM3024N'),
                                ('scc', 'MPPT Charge Controller 60', 'MC4860N15'),
                                ('network', 'RV-C network', 'rvc')]):
        devtype, friendly_name, modelname = modtup
        cursor = db.cursor()
        cursor.execute("insert into model (devtype, friendly_name, name) values (?, ?, ?);", (devtype, friendly_name, modelname))
        cursor.close()
        db.commit()

    enumids = set()

    # Populate the Params table and rvc_support at the same time
    enumids = fnPopulateParams(db, ws, xbdb)

    # Turn battery temperature compensation coefficient into a float instead
    # of an int (RVC database automatically turns this into an int because it
    # is a bit field
    fnTempCompConstIsFloat(db)

    # Populate the enum tables
    fnPopulateEnums(db, xbdb, enumids)

    # Turn DefaultBatteryTemperature into a customized enum
    fnDefaultBattTempEnum(db)

    # Same for Inverter Output Voltage, which is 108, 110 or 120 Vac
    fnInverterOutVoltageEnum(db)

    # InverterEnable is On/Off even though it is based on InvSts.Sts, which has multiple values
    fnInverterEnableEnum(db)

    # Turn On/Off for InvFaultRecovery into Auto/Manual
    fnInvFaultRecoveryEnum(db)

    # Turn LBCO Delay into an enum
    fnLBCODelayEnum(db)

    # Turn Power limit timer into an enum
    fnPwrLimitTimerEnum(db)

    # Turn Power Save timeout into an enum
    fnPwrSaveTimeEnum(db)

    # Populate the model support table
    for col, pattern in RVC_SUPPORT_LIST:
        cursor = db.cursor()
        cursor.execute(f"select id from model where friendly_name like '{pattern}';")
        while True:
            try:
                model_id = cursor.fetchone()[0]
            except:
                break

            fnPopulateModelSupport(db, ws, model_id, col)

        cursor.close()

    # Populate the limits for model support for the FEX4K
    fnUpdateModelLimits(db, 1)


    #--- Solar Charge Controller support ---
    for model_name in ['DR3210N', 'LS GM3024N']:
        model_id = fnGetModelId(db, name=model_name)
        fnSolarChargeControllerSupport(db, SCC_SPREADSHEET, model_id)

    # Strip out the Batt2 support from the PWM model
    model_id = fnGetModelId(db, name='LS GM3024N')
    cursor = db.cursor()
    cursor.execute("select id from param where name like 'Batt2%';")
    while True:
        try:
            param_id = cursor.fetchone()[0]
        except TypeError:
            break

        db.execute("delete from modbus_support where model_id = ? and param_id = ?",
                   (model_id, param_id))

    cursor.close()

    # Add in parameter descriptions for all three SCC models
    for model_name in ['DR3210N', 'LS GM3024N', 'MPPT_RVC_30']:
        model_id = fnGetModelId(db, name=model_name)
        fnSolarChargeDescriptions(db, SCC_SPREADSHEET, model_id)

    # Add in parameter descriptions for the FXC series of inverter/chargers and FX inverters
    # and the SRNE SCCs
    for friendly_pattern, spreadsheet in [('Freedom X%', FXCC_SPREADSHEET),
                                          ('MPPT Charge Controller 60', SRNE_SPREADSHEET)]:
        cursor = db.cursor()
        cursor.execute(f"select id from model where friendly_name like '{friendly_pattern}'")
        while True:
            try:
                model_id = cursor.fetchone()[0]
            except:
                break

            fnGetDescriptions(db, spreadsheet, model_id)

        cursor.close()

    # Add in parameter descriptions and cautions for the ZeroRPM BMS

    #--- Some hand edits not available from the settings spreadsheet ---

    # Customized/reduced enums for model support
    enum_filters = [("like '%AcInput0'", "(enum_val like 'Grid%' or enum_val like 'Shore%' or enum_val like '%Load%')"),
                    ("like '%AcInputOut0'", "(enum_val like 'Gen%' or enum_val like '%Load%')"),
                    ("like '%AcOut0'", "(enum_val like 'Grid%' or enum_val like 'Shore%' or enum_val like 'Gen%')"),
                    ("like '%OpMode'", "enum_val like 'Cbg%'"),
                    ("like 'FEX%Opmode'", "enum_val != 'Safe' and enum_val != 'Operating'"),
                    ("like '%_InverterSearchEnable'", "enum_val = 'Error' or enum_val = 'Unknown'"),
                    ("like '%_ChargerEnable'", "enum_val = 'Error'"),
                    ("like '%_AudibleAlarm'", "enum_val = 'Error'"),
                    ("like '%_InvFaultRecovery'", "enum_val = 'Error'"),
                    ("like '%_TransferMode'", "enum_val = 'Error'"),
                    ("like '%_InvIgnitionContrl'", "enum_val = 'Error'"),
                    ("like '%_BattSwitch'", "enum_val = 'Error'"),
                    ("like '%_ChargerSwitch'", "enum_val = 'Error'"),
                    ("like 'F%ChargerAlgorithm'", "enum_val not like '%Stage'"),
                    ("like '%_ChargerMode'", "enum_val like 'Linked%'"),
                    ("like 'MC4860N15_BatteryType'", "enum_val like '%Li-Ion%'"),
                    ("like 'MC4860N15_BatteryType'", "enum_val = 'Custom 1'", 'Custom'),
                    ("like '%_ChargerEnable'", "enum_ord > 1"),
                    ("like '%_ChargerEnable'", "enum_ord = 0", 'Off'),
                    ("like '%_ChargerEnable'", "enum_ord = 1", 'On'),
                    ]
    for filter_tup in enum_filters:
        if len(filter_tup) == 2:
            enum_selector, delete_where = filter_tup
        elif len(filter_tup) == 3:
            enum_selector, change_where, replacement = filter_tup
        else:
            raise Exception("enum_filters entry needs to have either 2 or 3 elements")
        
        sql1 = f"select id from enum where name {enum_selector};"
        cursor = db.cursor()
        cursor.execute(sql1)
        while True:
            try:
                enum_id = cursor.fetchone()[0]
            except:
                break

            c2 = db.cursor()
            if len(filter_tup) == 2:
                sql2 = f"delete from enum_values where enum_id = ? and {delete_where};"
                c2.execute(sql2, (enum_id,))
            elif len(filter_tup) == 3:
                sql2 = f"update enum_values set enum_val = ? where enum_id = ? and {change_where}"
                c2.execute(sql2, (replacement, enum_id))
            c2.close()

        cursor.close()

    # Erase enums for InverterLowBatteryCutoutDelayTime on FEX4K models and
    # replace with enums specific for their range (0 to 600 step 5)
    fnReplace4kLbcoDelay(db)

    # Erase enum for InverterOutVoltage on Freedom X 2000 230V and replace with enum specific to its
    # ratings (220, 230, 240)
    fnReplace230vInverterOutVoltage(db)

    # Hand-edits for PPN parameters, as they are not in the RVC database
    fnZeroRpmEdits(db)
    
    # Tweak limits for specific models
    limit_filters = [("like 'Freedom XC%3000'", "ChargerMaxChargeCurrent", "default_float = 150.0, max_float = 150.0"),
                     ("like 'Freedom XC%3000'", "ACBreakerRating", "default_float = 50.0, max_float = 50.0"),
                     ("like 'Freedom XC%3000'", "InverterOutPowerLimit", "default_float = 3000.0, max_float = 3000.0"),
                     ("= 'Freedom XC 2000'", "ChargerMaxChargeCurrent", "default_float = 80.0, max_float = 80.0"),
                     ("= 'Freedom XC 1000'", "ChargerMaxChargeCurrent", "default_float = 50.0, max_float = 50.0"),
                     ("= 'Freedom XC 1000'", "InverterOutPowerLimit", "default_float = 1000.0, max_float = 1000.0"),
                     ("= 'Freedom X 1000'", "InverterOutPowerLimit", "default_float = 1000.0, max_float = 1000.0"),
                     ("= 'Freedom X 1200'", "InverterOutPowerLimit", "default_float = 1200.0, max_float = 1200.0"),
                     ("= 'Freedom X 3000'", "InverterOutPowerLimit", "default_float = 3000.0, max_float = 3000.0"),
                     ("= 'Freedom X 2000 230V'", "InverterOutVoltage", "default_float = 230.0, min_float = 220.0, max_float = 240.0, step_float=10.0"),
                     ("= 'Freedom X 2000 230V'", "InverterOutFrequency", "default_float = 50.0"),
                     ("= 'Freedom X 2000 230V'", "AcUvSdThresh", "default_float = 180.0, min_float = 170.0, max_float = 220.0"),
                     ("= 'Freedom X 2000 24V'", "LbcoRecovThresh", "default_float = 26.2, min_float = 20.4, max_float = 32.0"),
                     ("= 'Freedom X 2000 24V'", "InverterLowBatteryCutoutVoltage", "default_float = 21.0, min_float = 20.0, max_float = 25.6")]
    for model_clause, param_name, set_clause in limit_filters:
        param_id = fnGetParamId(db, name=param_name)

        c1 = db.cursor()
        c1.execute("select id from model where friendly_name %s;" % model_clause)
        while True:
            try:
                model_id = c1.fetchone()[0]
            except:
                break

            db.execute("update model_support set %s where model_id = ? and param_id = ?" % set_clause, (model_id, param_id))

        c1.close()

    # Close and exit
    db.commit()
    db.close()

if __name__ == '__main__':
    fnMain()
